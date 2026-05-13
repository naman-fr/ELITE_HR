import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
OUTPUT_FILE = "ELITE_HR_Master_Dashboard.xlsx"
NUM_INDIA_EMP = 30
NUM_US_EMP = 10

def create_mock_data():
    # India Database
    india_data = {
        "Employee ID": [f"IN10{i:02d}" for i in range(1, NUM_INDIA_EMP + 1)],
        "Name": [f"Employee {i}" for i in range(1, NUM_INDIA_EMP + 1)],
        "Department": np.random.choice(["Engineering", "HR", "Sales", "Finance", "Product", "Operations"], NUM_INDIA_EMP),
        "Region": ["India"] * NUM_INDIA_EMP,
        "DOJ": pd.to_datetime(np.random.choice(pd.date_range("2020-01-01", "2024-01-01"), NUM_INDIA_EMP)),
        "Reporting Manager": np.random.choice(["Manager A", "Manager B", "Manager C"], NUM_INDIA_EMP),
        "Employment Status": np.random.choice(["Confirmed", "Under Probation", "Intern"], NUM_INDIA_EMP, p=[0.7, 0.2, 0.1]),
        "Email": [f"emp{i}@company.com" for i in range(1, NUM_INDIA_EMP + 1)],
        "Phone": [f"+91 98765 432{i:02d}" for i in range(1, NUM_INDIA_EMP + 1)],
    }
    df_india = pd.DataFrame(india_data)

    # US Database
    us_data = {
        "Employee ID": [f"US20{i:02d}" for i in range(1, NUM_US_EMP + 1)],
        "Name": [f"US Employee {i}" for i in range(1, NUM_US_EMP + 1)],
        "Department": np.random.choice(["Engineering", "Product", "Marketing"], NUM_US_EMP),
        "Region": ["US"] * NUM_US_EMP,
        "DOJ": pd.to_datetime(np.random.choice(pd.date_range("2018-01-01", "2023-01-01"), NUM_US_EMP)),
        "Reporting Manager": ["Director X"] * NUM_US_EMP,
        "Employment Status": ["Confirmed"] * NUM_US_EMP,
        "Email": [f"us_emp{i}@company.com" for i in range(1, NUM_US_EMP + 1)],
    }
    df_us = pd.DataFrame(us_data)

    # Finance
    finance_data = {
        "Employee ID": list(df_india["Employee ID"]) + list(df_us["Employee ID"]),
        "Currency": ["INR"] * NUM_INDIA_EMP + ["USD"] * NUM_US_EMP,
        "Annual CTC": [np.random.randint(600000, 3000000) for _ in range(NUM_INDIA_EMP)] + [np.random.randint(80000, 180000) for _ in range(NUM_US_EMP)],
    }
    df_finance = pd.DataFrame(finance_data)

    # Productivity
    prod_data = {
        "Employee ID": list(df_india["Employee ID"]) + list(df_us["Employee ID"]),
        "Avg Hrs/Day": [np.random.uniform(6.5, 9.5) for _ in range(NUM_INDIA_EMP + NUM_US_EMP)],
    }
    df_prod = pd.DataFrame(prod_data)
    df_prod["Flag"] = df_prod["Avg Hrs/Day"].apply(lambda x: "⚠ Below 8 Hrs" if x < 8 else "✅ OK")

    return df_india, df_us, df_finance, df_prod

def apply_styling(ws, title):
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column width
        ws.column_dimensions[get_column_letter(col)].width = 20

def generate():
    df_india, df_us, df_finance, df_prod = create_mock_data()
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_india.to_excel(writer, sheet_name="India Employee Database", index=False)
        df_us.to_excel(writer, sheet_name="US Employee Database", index=False)
        df_finance.to_excel(writer, sheet_name="Finance", index=False)
        df_prod.to_excel(writer, sheet_name="Productivity", index=False)
        
        # Create other tabs
        for tab in ["Dashboard", "Risk Report", "RM Data", "Offboarded Resources", "SecOps_Wazuh", "SecOps_Wazuh_DLP", "SecOps_Keycloak", "CALC", "CONFIG", "AI_CONTEXT"]:
            pd.DataFrame().to_excel(writer, sheet_name=tab, index=False)

    # Apply styling to all sheets
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_column > 1:
            apply_styling(ws, sheet)
    
    wb.save(OUTPUT_FILE)
    print(f"Excel file created: {OUTPUT_FILE}")

if __name__ == "__main__":
    generate()
